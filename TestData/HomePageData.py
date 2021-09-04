import openpyxl

class HomePageData:

    test_HomePage_data = [{"firstname":"Alex", "email":"a_vazquez_d@yahoo.com", "password":"12345", "gender":"Male", "type":"Student"},
                            {"firstname":"Alma", "email":"almaDd@yahoo.com", "password":"56789", "gender":"Female", "type":"Employed"},
                            {"firstname":"Hector", "email":"hector@yahoo.com", "password":"45680", "gender":"Humane", "type":"Student"}]

    @staticmethod
    def getTestData(test_case_name):

        book = openpyxl.load_workbook("D:\\DEV\\PYTHON\\xlsDemo.xlsx")
        sheet = book.active
        cell = sheet.cell(row=1, column=2)
        test = 13213
        print(cell.value)
        dict = {}
        for i in range(1, sheet.max_row + 1):
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):
                    dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

        return [dict]
