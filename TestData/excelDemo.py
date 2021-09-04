import openpyxl
##### READ
book = openpyxl.load_workbook("D:\\DEV\\PYTHON\\xlsDemo.xlsx")
sheet = book.active
cell = sheet.cell(row=1, column=2)
print(cell.value)

dict = {}

for i in range(1, sheet.max_row+1):
    for j in range(2, sheet.max_column+1):
        dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value

